import csv
import json
import openpyxl
from decimal import Decimal
from datetime import date as date_cls, timedelta
from datetime import datetime as dt_class
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.cache import cache
from django.db.models import Sum, Count, Min, Max, Q, Case, When, F, DecimalField, ExpressionWrapper, Value, Subquery, OuterRef
from django.db.models.functions import TruncMonth, TruncWeek, ExtractYear, ExtractMonth
from django.db.models.functions import Coalesce
from django.conf import settings as django_settings
from django.http import HttpResponse, JsonResponse
from django.utils import timezone

from .models import Distributor, POSUpload, POSRecord, ExchangeRate, MonthlyRate, PriorityProduct
from .forms import UploadForm
from .parsers import get_parser

DIST_COLORS = [
    '#8205B4', '#0EA5E9', '#F59E0B', '#10B981', '#EF4444', '#6366F1',
    '#EC4899', '#14B8A6', '#F97316', '#84CC16',
]

REGION_COLORS = {
    'APAC':     '#0EA5E9',
    'EMEA':     '#8205B4',
    'Americas': '#10B981',
}


# ── Currency helpers ───────────────────────────────────────────────────────────

def _get_rates():
    """Return dict {currency: {usd: Decimal, eur: Decimal}}. Cached 1 h; refreshes from API if DB rates are >24 h old."""
    cached = cache.get('kpos_get_rates')
    if cached is not None:
        return cached

    from datetime import timedelta
    threshold = timezone.now() - timedelta(hours=24)
    if not ExchangeRate.objects.filter(fetched_at__gte=threshold).exists():
        try:
            import urllib.request
            import json as _json
            url = 'https://open.er-api.com/v6/latest/USD'
            with urllib.request.urlopen(url, timeout=3) as r:
                data = _json.loads(r.read())
            api_rates = data['rates']  # {EUR: X, GBP: X, SEK: X, DKK: X} all per 1 USD
            eur_per_usd = Decimal(str(api_rates['EUR']))
            entries = {
                'USD': (Decimal('1'), eur_per_usd),
                'EUR': (Decimal('1') / eur_per_usd, Decimal('1')),
                'GBP': (Decimal('1') / Decimal(str(api_rates['GBP'])),
                        eur_per_usd / Decimal(str(api_rates['GBP']))),
                'SEK': (Decimal('1') / Decimal(str(api_rates['SEK'])),
                        eur_per_usd / Decimal(str(api_rates['SEK']))),
                'DKK': (Decimal('1') / Decimal(str(api_rates['DKK'])),
                        eur_per_usd / Decimal(str(api_rates['DKK']))),
            }
            for cur, (r_usd, r_eur) in entries.items():
                ExchangeRate.objects.update_or_create(
                    currency=cur,
                    defaults={'rate_to_usd': r_usd, 'rate_to_eur': r_eur},
                )
            # Warm the _annotate_converted cache so it doesn't re-query the DB
            cache.set('kpos_current_rates', {
                cur: (float(r_usd), float(r_eur))
                for cur, (r_usd, r_eur) in entries.items()
            }, 3600)
        except Exception:
            pass

    result = {r.currency: {'usd': r.rate_to_usd, 'eur': r.rate_to_eur}
              for r in ExchangeRate.objects.all()}
    cache.set('kpos_get_rates', result, 3600)  # Cache for 1 h — no DB hit on next 9 page views
    return result


def _annotate_converted(qs, target, rates=None):
    """Annotate queryset with converted_value using a CASE/WHEN built from cached rates.

    Replaces correlated subqueries (one per row) with a single CASE expression
    compiled from preloaded monthly + current rates — dramatically faster on large tables.
    """
    field_idx = 0 if target == 'USD' else 1

    # Load monthly rates from cache (small table, rarely changes)
    monthly_rates = cache.get('kpos_monthly_rates')
    if monthly_rates is None:
        monthly_rates = {
            (r.year, r.month, r.currency): (float(r.rate_to_usd), float(r.rate_to_eur))
            for r in MonthlyRate.objects.all()
        }
        cache.set('kpos_monthly_rates', monthly_rates, 86400)

    current_rates = cache.get('kpos_current_rates')
    if current_rates is None:
        current_rates = {
            r.currency: (float(r.rate_to_usd), float(r.rate_to_eur))
            for r in ExchangeRate.objects.all()
        }
        cache.set('kpos_current_rates', current_rates, 3600)

    _df = DecimalField(max_digits=12, decimal_places=6)
    cases = []
    for (year, month, currency), rate_pair in monthly_rates.items():
        cases.append(When(
            invoice_date__year=year, invoice_date__month=month, currency=currency,
            then=Value(Decimal(str(rate_pair[field_idx])), output_field=_df),
        ))
    for currency, rate_pair in current_rates.items():
        cases.append(When(
            currency=currency,
            then=Value(Decimal(str(rate_pair[field_idx])), output_field=_df),
        ))

    rate_expr = Case(*cases, default=Value(Decimal('1'), output_field=_df), output_field=_df)

    return qs.annotate(
        converted_value=ExpressionWrapper(
            F('invoiced_value') * rate_expr,
            output_field=DecimalField(max_digits=15, decimal_places=4),
        )
    )


def _currency_symbol(currency):
    return '€' if currency == 'EUR' else '$'


def set_currency(request):
    """POST endpoint to set session currency preference."""
    cur = request.POST.get('currency', 'USD')
    if cur in ('USD', 'EUR'):
        request.session['currency'] = cur
    return redirect(request.POST.get('next', '/'))


def set_region(request):
    """POST endpoint to set session region filter."""
    region = request.POST.get('region', '')
    request.session['region'] = region
    request.session.modified = True
    return redirect(request.POST.get('next', '/'))


# ── Prior-period comparison helper ────────────────────────────────────────────

def _prior_period_stats(qs_base, date_from_str, date_to_str):
    """Compute total revenue+units for the period immediately before the current window."""
    today = date_cls.today()
    try:
        if date_from_str and date_to_str:
            df = dt_class.strptime(date_from_str, '%Y-%m-%d').date()
            dt = dt_class.strptime(date_to_str, '%Y-%m-%d').date()
        elif date_from_str:
            df = dt_class.strptime(date_from_str, '%Y-%m-%d').date()
            dt = today
        elif date_to_str:
            dt = dt_class.strptime(date_to_str, '%Y-%m-%d').date()
            df = dt - timedelta(days=29)
        else:
            df = today.replace(day=1)
            dt = today
        n_days = max((dt - df).days + 1, 1)
        prior_dt = df - timedelta(days=1)
        prior_df = prior_dt - timedelta(days=n_days - 1)
    except (ValueError, OverflowError):
        return None
    result = qs_base.filter(
        invoice_date__gte=prior_df,
        invoice_date__lte=prior_dt,
    ).aggregate(revenue=Sum('invoiced_value'), units=Sum('quantity'))
    return {
        'revenue': float(result['revenue'] or 0),
        'units': int(result['units'] or 0),
        'label': prior_df.strftime('%b %d') + ' – ' + prior_dt.strftime('%b %d, %Y'),
    }


# ── ISO country map ────────────────────────────────────────────────────────────

ISO_COUNTRIES = {
    'AF': 'Afghanistan', 'AX': 'Åland Islands', 'AL': 'Albania', 'DZ': 'Algeria',
    'AS': 'American Samoa', 'AD': 'Andorra', 'AO': 'Angola', 'AI': 'Anguilla',
    'AQ': 'Antarctica', 'AG': 'Antigua & Barbuda', 'AR': 'Argentina', 'AM': 'Armenia',
    'AW': 'Aruba', 'AU': 'Australia', 'AT': 'Austria', 'AZ': 'Azerbaijan',
    'BS': 'Bahamas', 'BH': 'Bahrain', 'BD': 'Bangladesh', 'BB': 'Barbados',
    'BY': 'Belarus', 'BE': 'Belgium', 'BZ': 'Belize', 'BJ': 'Benin',
    'BM': 'Bermuda', 'BT': 'Bhutan', 'BO': 'Bolivia', 'BA': 'Bosnia & Herzegovina',
    'BW': 'Botswana', 'BR': 'Brazil', 'BN': 'Brunei', 'BG': 'Bulgaria',
    'BF': 'Burkina Faso', 'BI': 'Burundi', 'CV': 'Cabo Verde', 'KH': 'Cambodia',
    'CM': 'Cameroon', 'CA': 'Canada', 'KY': 'Cayman Islands', 'CF': 'Central African Republic',
    'TD': 'Chad', 'CL': 'Chile', 'CN': 'China', 'CX': 'Christmas Island',
    'CC': 'Cocos Islands', 'CO': 'Colombia', 'KM': 'Comoros', 'CG': 'Congo',
    'CD': 'DR Congo', 'CK': 'Cook Islands', 'CR': 'Costa Rica', 'CI': "Côte d'Ivoire",
    'HR': 'Croatia', 'CU': 'Cuba', 'CW': 'Curaçao', 'CY': 'Cyprus',
    'CZ': 'Czech Republic', 'DK': 'Denmark', 'DJ': 'Djibouti', 'DM': 'Dominica',
    'DO': 'Dominican Republic', 'EC': 'Ecuador', 'EG': 'Egypt', 'SV': 'El Salvador',
    'GQ': 'Equatorial Guinea', 'ER': 'Eritrea', 'EE': 'Estonia', 'SZ': 'Eswatini',
    'ET': 'Ethiopia', 'FK': 'Falkland Islands', 'FO': 'Faroe Islands', 'FJ': 'Fiji',
    'FI': 'Finland', 'FR': 'France', 'GF': 'French Guiana', 'PF': 'French Polynesia',
    'GA': 'Gabon', 'GM': 'Gambia', 'GE': 'Georgia', 'DE': 'Germany',
    'GH': 'Ghana', 'GI': 'Gibraltar', 'GR': 'Greece', 'GL': 'Greenland',
    'GD': 'Grenada', 'GP': 'Guadeloupe', 'GU': 'Guam', 'GT': 'Guatemala',
    'GG': 'Guernsey', 'GN': 'Guinea', 'GW': 'Guinea-Bissau', 'GY': 'Guyana',
    'HT': 'Haiti', 'HN': 'Honduras', 'HK': 'Hong Kong', 'HU': 'Hungary',
    'IS': 'Iceland', 'IN': 'India', 'ID': 'Indonesia', 'IR': 'Iran',
    'IQ': 'Iraq', 'IE': 'Ireland', 'IM': 'Isle of Man', 'IL': 'Israel',
    'IT': 'Italy', 'JM': 'Jamaica', 'JP': 'Japan', 'JE': 'Jersey',
    'JO': 'Jordan', 'KZ': 'Kazakhstan', 'KE': 'Kenya', 'KI': 'Kiribati',
    'KP': 'North Korea', 'KR': 'South Korea', 'KW': 'Kuwait', 'KG': 'Kyrgyzstan',
    'LA': 'Laos', 'LV': 'Latvia', 'LB': 'Lebanon', 'LS': 'Lesotho',
    'LR': 'Liberia', 'LY': 'Libya', 'LI': 'Liechtenstein', 'LT': 'Lithuania',
    'LU': 'Luxembourg', 'MO': 'Macao', 'MG': 'Madagascar', 'MW': 'Malawi',
    'MY': 'Malaysia', 'MV': 'Maldives', 'ML': 'Mali', 'MT': 'Malta',
    'MH': 'Marshall Islands', 'MQ': 'Martinique', 'MR': 'Mauritania', 'MU': 'Mauritius',
    'YT': 'Mayotte', 'MX': 'Mexico', 'FM': 'Micronesia', 'MD': 'Moldova',
    'MC': 'Monaco', 'MN': 'Mongolia', 'ME': 'Montenegro', 'MS': 'Montserrat',
    'MA': 'Morocco', 'MZ': 'Mozambique', 'MM': 'Myanmar', 'NA': 'Namibia',
    'NR': 'Nauru', 'NP': 'Nepal', 'NL': 'Netherlands', 'NC': 'New Caledonia',
    'NZ': 'New Zealand', 'NI': 'Nicaragua', 'NE': 'Niger', 'NG': 'Nigeria',
    'NU': 'Niue', 'NF': 'Norfolk Island', 'MK': 'North Macedonia', 'MP': 'Northern Mariana Islands',
    'NO': 'Norway', 'OM': 'Oman', 'PK': 'Pakistan', 'PW': 'Palau',
    'PS': 'Palestine', 'PA': 'Panama', 'PG': 'Papua New Guinea', 'PY': 'Paraguay',
    'PE': 'Peru', 'PH': 'Philippines', 'PN': 'Pitcairn', 'PL': 'Poland',
    'PT': 'Portugal', 'PR': 'Puerto Rico', 'QA': 'Qatar', 'RE': 'Réunion',
    'RO': 'Romania', 'RU': 'Russia', 'RW': 'Rwanda', 'BL': 'Saint Barthélemy',
    'SH': 'Saint Helena', 'KN': 'Saint Kitts & Nevis', 'LC': 'Saint Lucia',
    'MF': 'Saint Martin', 'PM': 'Saint Pierre & Miquelon', 'VC': 'Saint Vincent',
    'WS': 'Samoa', 'SM': 'San Marino', 'ST': 'São Tomé & Príncipe', 'SA': 'Saudi Arabia',
    'SN': 'Senegal', 'RS': 'Serbia', 'SC': 'Seychelles', 'SL': 'Sierra Leone',
    'SG': 'Singapore', 'SX': 'Sint Maarten', 'SK': 'Slovakia', 'SI': 'Slovenia',
    'SB': 'Solomon Islands', 'SO': 'Somalia', 'ZA': 'South Africa', 'GS': 'South Georgia',
    'SS': 'South Sudan', 'ES': 'Spain', 'LK': 'Sri Lanka', 'SD': 'Sudan',
    'SR': 'Suriname', 'SJ': 'Svalbard & Jan Mayen', 'SE': 'Sweden', 'CH': 'Switzerland',
    'SY': 'Syria', 'TW': 'Taiwan', 'TJ': 'Tajikistan', 'TZ': 'Tanzania',
    'TH': 'Thailand', 'TL': 'Timor-Leste', 'TG': 'Togo', 'TK': 'Tokelau',
    'TO': 'Tonga', 'TT': 'Trinidad & Tobago', 'TN': 'Tunisia', 'TR': 'Turkey',
    'TM': 'Turkmenistan', 'TC': 'Turks & Caicos Islands', 'TV': 'Tuvalu',
    'UG': 'Uganda', 'UA': 'Ukraine', 'AE': 'United Arab Emirates', 'GB': 'United Kingdom',
    'US': 'United States', 'UY': 'Uruguay', 'UZ': 'Uzbekistan', 'VU': 'Vanuatu',
    'VE': 'Venezuela', 'VN': 'Vietnam', 'VG': 'British Virgin Islands',
    'VI': 'US Virgin Islands', 'WF': 'Wallis & Futuna', 'EH': 'Western Sahara',
    'YE': 'Yemen', 'ZM': 'Zambia', 'ZW': 'Zimbabwe',
}


def normalize_country(raw):
    """Normalize any country value (ISO code, all-caps, mixed case) to a canonical full name."""
    if not raw:
        return raw
    raw = raw.strip()
    if not raw:
        return raw
    upper = raw.upper()
    # 2-letter ISO code
    if len(raw) == 2 and upper in ISO_COUNTRIES:
        return ISO_COUNTRIES[upper]
    # Match against known full names (case-insensitive)
    lower = raw.lower()
    for name in ISO_COUNTRIES.values():
        if name.lower() == lower:
            return name
    # Fallback: title-case if all-caps, otherwise keep as-is
    return raw.title() if raw.isupper() else raw


def country_display(raw):
    return normalize_country(raw)


# ── Views ──────────────────────────────────────────────────────────────────────

def dashboard(request):
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    distributor_id = request.GET.get('distributor', '')

    # Region: explicit URL param updates session; navigation falls back to session
    region_param = request.GET.get('region')
    if region_param is not None:
        region = region_param.strip()
        request.session['region'] = region
        request.session.modified = True
    else:
        region = request.session.get('region', '')

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    # Base queryset without date filters (for comparison)
    qs_base = POSRecord.objects.select_related('distributor')
    if region:
        qs_base = qs_base.filter(distributor__region=region)
    if distributor_id:
        qs_base = qs_base.filter(distributor_id=distributor_id)

    # Current-period queryset
    qs = qs_base
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)

    qs_conv = _annotate_converted(qs, selected_currency, rates)
    totals = qs_conv.aggregate(
        total_revenue=Sum('converted_value'),
        total_units=Sum('quantity'),
        unique_countries=Count('country', distinct=True),
    )
    active_distributors = qs.values('distributor').distinct().count()

    # Prior-period comparison (native currency, approximate)
    prior = _prior_period_stats(qs_base, date_from, date_to)
    current_rev   = float(totals['total_revenue'] or 0)
    current_units = int(totals['total_units'] or 0)
    rev_change_pct   = None
    units_change_pct = None
    if prior:
        if prior['revenue'] > 0:
            rev_change_pct = round((current_rev - prior['revenue']) / prior['revenue'] * 100, 1)
        if prior['units'] > 0:
            units_change_pct = round((current_units - prior['units']) / prior['units'] * 100, 1)
    today = date_cls.today()
    if date_from or date_to:
        prior_label = 'vs prev period'
    else:
        first_of_month = today.replace(day=1)
        prior_month_last = first_of_month - timedelta(days=1)
        prior_label = 'vs ' + prior_month_last.strftime('%b %Y')

    # Monthly revenue by region for wave chart
    monthly_qs = list(
        _annotate_converted(
            qs.filter(invoice_date__isnull=False, invoiced_value__isnull=False),
            selected_currency, rates
        )
        .annotate(month=TruncMonth('invoice_date'))
        .values('month', 'distributor__region')
        .annotate(revenue=Sum('converted_value'))
        .order_by('month', 'distributor__region')
    )
    months_sorted = sorted(set(r['month'] for r in monthly_qs))
    month_labels  = [m.strftime('%b %Y') for m in months_sorted]
    month_keys    = [m.strftime('%Y-%m') for m in months_sorted]

    region_order = {}
    for r in monthly_qs:
        reg = r['distributor__region'] or 'Unknown'
        if reg not in region_order:
            region_order[reg] = {'by_month': {}}
        region_order[reg]['by_month'][r['month'].strftime('%Y-%m')] = float(r['revenue'])

    datasets = []
    for reg, info in region_order.items():
        color = REGION_COLORS.get(reg, '#6c757d')
        datasets.append({
            'label': reg,
            'data': [info['by_month'].get(mk, 0) for mk in month_keys],
            'backgroundColor': color,
            'borderColor': color,
            'borderWidth': 0,
            'fill': False,
            'tension': 0,
            'pointRadius': 0,
            'pointHoverRadius': 0,
            'borderRadius': 4,
            '_color': color,
            'region': reg,
        })
    chart_data = json.dumps({'labels': month_labels, 'datasets': datasets})

    # Region cards + donut chart data
    region_donut_raw = list(
        _annotate_converted(
            qs.filter(invoiced_value__isnull=False),
            selected_currency, rates
        )
        .values('distributor__region')
        .annotate(
            revenue=Sum('converted_value'),
            records=Count('id'),
            dist_count=Count('distributor', distinct=True),
            country_count=Count('country', distinct=True),
        )
        .order_by('-revenue')
    )
    total_donut = sum(float(r['revenue'] or 0) for r in region_donut_raw) or 1
    region_cards = [
        {
            'name':       r['distributor__region'] or 'Unknown',
            'revenue':    float(r['revenue'] or 0),
            'records':    r['records'] or 0,
            'dist_count': r['dist_count'] or 0,
            'country_count': r['country_count'] or 0,
            'color':      REGION_COLORS.get(r['distributor__region'], '#6c757d'),
            'pct':        round(float(r['revenue'] or 0) / total_donut * 100, 1),
        }
        for r in region_donut_raw
    ]
    region_donut = json.dumps({
        'labels': [r['name'] for r in region_cards],
        'data':   [round(r['revenue'], 2) for r in region_cards],
        'colors': [r['color'] for r in region_cards],
        'pcts':   [r['pct'] for r in region_cards],
    })

    # Top 10 products by Kramer part number
    top_products = list(
        _annotate_converted(
            qs.exclude(manufacturer_part_no=''),
            selected_currency, rates
        )
        .values('manufacturer_part_no')
        .annotate(
            description=Max('product_description'),
            revenue=Sum('converted_value'),
            units=Sum('quantity'),
        )
        .order_by('-revenue')[:10]
    )
    if top_products:
        max_rev = float(max(p['revenue'] for p in top_products) or 1)
        for p in top_products:
            p['revenue'] = float(p['revenue'] or 0)
            p['pct'] = round(p['revenue'] / max_rev * 100)

    # Top 5 account managers
    top_salespersons = list(
        _annotate_converted(
            qs.exclude(distributor__salesperson_name='').filter(invoiced_value__isnull=False),
            selected_currency, rates
        )
        .values('distributor__salesperson_name', 'distributor__salesperson_code')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'), dist_count=Count('distributor', distinct=True))
        .order_by('-revenue')[:5]
    )
    if top_salespersons:
        max_sp_rev = float(max(sp['revenue'] for sp in top_salespersons) or 1)
        for sp in top_salespersons:
            sp['revenue'] = float(sp['revenue'] or 0)
            sp['units'] = int(sp['units'] or 0)
            sp['pct'] = round(sp['revenue'] / max_sp_rev * 100)
    else:
        for sp in top_salespersons:
            sp['revenue'] = float(sp['revenue'] or 0)
            sp['units'] = int(sp['units'] or 0)
            sp['pct'] = 0

    # Per-distributor summary table
    dist_summary = list(
        _annotate_converted(qs, selected_currency, rates)
        .values('distributor__id', 'distributor__name', 'distributor__region', 'distributor__code')
        .annotate(
            revenue=Sum('converted_value'),
            units=Sum('quantity'),
            countries=Count('country', distinct=True),
            records=Count('id'),
        )
        .order_by('-revenue')
    )
    total_dist_rev = sum(float(d['revenue'] or 0) for d in dist_summary) or 1
    for d in dist_summary:
        d['revenue_f']  = float(d['revenue'] or 0)
        d['share_pct']  = round(d['revenue_f'] / total_dist_rev * 100, 1)

    all_regions      = Distributor.objects.exclude(region='').values_list('region', flat=True).distinct().order_by('region')
    all_distributors = (Distributor.objects.filter(region=region) if region else Distributor.objects.all()).order_by('name')

    # User greeting
    if request.user.is_authenticated:
        full_name = request.user.get_full_name()
        user_display = full_name or request.user.username
    else:
        user_display = ''
    name_parts = user_display.split() if user_display else []
    initials = ((name_parts[0][0] if name_parts else '') +
                (name_parts[-1][0] if len(name_parts) > 1 else '')).upper() or 'K'

    context = {
        'total_revenue':      current_rev,
        'total_units':        current_units,
        'active_distributors': active_distributors,
        'unique_countries':   totals['unique_countries'] or 0,
        'rev_change_pct':     rev_change_pct,
        'units_change_pct':   units_change_pct,
        'prior_label':        prior_label,
        'chart_data':         chart_data,
        'region_donut':       region_donut,
        'dist_summary':       dist_summary,
        'all_regions':        all_regions,
        'all_distributors':   all_distributors,
        'filters': {
            'date_from':    date_from,
            'date_to':      date_to,
            'region':       region,
            'distributor':  distributor_id,
        },
        'top_products':      top_products,
        'top_salespersons':  top_salespersons,
        'region_cards':   region_cards,
        'page_title':     'Revenue Dashboard',
        'has_filters':    any([date_from, date_to, distributor_id]),
        'today_date':     today,
        'user_display':   user_display,
        'initials':       initials,
        'selected_currency': selected_currency,
        'currency_symbol':   currency_symbol,
    }
    return render(request, 'reports/dashboard.html', context)


def distributor_records(request, pk):
    distributor = get_object_or_404(Distributor, pk=pk)
    records = POSRecord.objects.filter(distributor=distributor).select_related('upload')

    # Filters
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    category = request.GET.get('category', '').strip()
    upload_id = request.GET.get('upload', '').strip()

    if q:
        records = records.filter(
            Q(customer_name__icontains=q) |
            Q(product_name__icontains=q) |
            Q(item_number__icontains=q) |
            Q(order_ref__icontains=q) |
            Q(address_city__icontains=q)
        )
    if date_from:
        records = records.filter(invoice_date__gte=date_from)
    if date_to:
        records = records.filter(invoice_date__lte=date_to)
    if category:
        records = records.filter(product_level_1=category)
    if upload_id:
        records = records.filter(upload_id=upload_id)

    # Summary stats for filtered results
    stats = records.aggregate(
        total_records=Count('id'),
        total_qty=Sum('quantity'),
        total_value=Sum('invoiced_value'),
        unique_customers=Count('customer_name', distinct=True),
        date_from=Min('invoice_date'),
        date_to=Max('invoice_date'),
    )

    # Filter options
    categories = POSRecord.objects.filter(distributor=distributor).values_list(
        'product_level_1', flat=True).distinct().exclude(product_level_1='').order_by('product_level_1')
    uploads = distributor.uploads.all()

    context = {
        'distributor': distributor,
        'records': records,
        'stats': stats,
        'categories': categories,
        'uploads': uploads,
        'filters': {
            'q': q,
            'date_from': date_from,
            'date_to': date_to,
            'category': category,
            'upload': upload_id,
        },
        'page_title': distributor.name,
        'salesperson_code': distributor.salesperson_code,
        'salesperson_name': distributor.salesperson_name,
    }
    return render(request, 'reports/records.html', context)


def upload_file(request):
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            distributor = form.cleaned_data['distributor']
            excel_file = request.FILES['excel_file']
            report_period = form.cleaned_data['report_period']
            replace_existing = form.cleaned_data['replace_existing']
            notes = form.cleaned_data['notes']

            parser = get_parser(distributor.code)
            if not parser:
                messages.error(request, f'No parser configured for distributor "{distributor.name}" (code: {distributor.code}). Contact the system administrator.')
                return render(request, 'reports/upload.html', {'form': form, 'page_title': 'Upload Report'})

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                parsed_rows = parser(wb)
            except Exception as e:
                messages.error(request, f'Failed to parse Excel file: {e}')
                return render(request, 'reports/upload.html', {'form': form, 'page_title': 'Upload Report'})

            if not parsed_rows:
                messages.error(request, 'No data rows found in the uploaded file.')
                return render(request, 'reports/upload.html', {'form': form, 'page_title': 'Upload Report'})

            if replace_existing:
                POSRecord.objects.filter(distributor=distributor).delete()
                distributor.uploads.all().delete()

            upload = POSUpload.objects.create(
                distributor=distributor,
                original_filename=excel_file.name,
                report_period=report_period,
                row_count=len(parsed_rows),
                notes=notes,
            )

            bulk_records = [
                POSRecord(upload=upload, distributor=distributor, **row)
                for row in parsed_rows
            ]
            POSRecord.objects.bulk_create(bulk_records)

            try:
                from django.core.management import call_command
                call_command('update_monthly_rates', verbosity=0)
            except Exception:
                pass

            messages.success(request, f'Successfully imported {len(parsed_rows)} records from "{excel_file.name}".')
            return redirect('distributor_records', pk=distributor.pk)
    else:
        form = UploadForm()

    return render(request, 'reports/upload.html', {'form': form, 'page_title': 'Upload Report'})


def export_csv(request, pk):
    distributor = get_object_or_404(Distributor, pk=pk)
    records = POSRecord.objects.filter(distributor=distributor).order_by('-invoice_date')

    # Apply same filters as records view
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    category = request.GET.get('category', '').strip()

    if q:
        records = records.filter(
            Q(customer_name__icontains=q) |
            Q(product_name__icontains=q) |
            Q(item_number__icontains=q) |
            Q(order_ref__icontains=q)
        )
    if date_from:
        records = records.filter(invoice_date__gte=date_from)
    if date_to:
        records = records.filter(invoice_date__lte=date_to)
    if category:
        records = records.filter(product_level_1=category)

    filename = f"POS_{distributor.code}_{timezone.now().strftime('%Y%m%d')}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'Invoice Date', 'Invoice Ref', 'Order Ref',
        'Customer Account', 'Customer Name', 'City', 'Country',
        'Item Number', 'Product Name', 'Manufacturer Part No', 'Description',
        'Category L1', 'Category L2', 'Category L3',
        'Brand', 'Quantity', 'Sales Price', 'Invoiced Value', 'Currency',
        'Vendor', 'SDA Number', 'Special Bid',
        'Address', 'Post Code', 'Telephone',
    ])
    for r in records:
        writer.writerow([
            r.invoice_date, r.invoice_ref, r.order_ref,
            r.customer_account, r.customer_name, r.address_city, r.country,
            r.item_number, r.product_name, r.manufacturer_part_no, r.product_description,
            r.product_level_1, r.product_level_2, r.product_level_3,
            r.brand, r.quantity, r.sales_price, r.invoiced_value, r.currency,
            r.vendor, r.sda_number, r.special_bid_number,
            r.address_street, r.post_code, r.telephone,
        ])

    return response


def distributor_list(request):
    # Region: explicit URL param updates session; navigation falls back to session
    region_param = request.GET.get('region')
    if region_param is not None:
        selected_region = region_param.strip()
        request.session['region'] = selected_region
        request.session.modified = True
    else:
        selected_region = request.session.get('region', '')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    selected_sp = request.GET.get('salesperson', '').strip()

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    all_distributors = Distributor.objects.all().order_by('name')
    jump_distributors = (
        Distributor.objects.filter(region=selected_region).order_by('name')
        if selected_region else all_distributors
    )

    # Region stats (always unfiltered by region)
    region_base = POSRecord.objects.filter(invoiced_value__isnull=False)
    if date_from:
        region_base = region_base.filter(invoice_date__gte=date_from)
    if date_to:
        region_base = region_base.filter(invoice_date__lte=date_to)

    region_qs = (
        _annotate_converted(region_base, selected_currency, rates)
        .values('distributor__region')
        .annotate(
            revenue=Sum('converted_value'),
            records=Count('id'),
            dist_count=Count('distributor', distinct=True),
        )
        .order_by('-revenue')
    )
    region_stats = [
        {
            'name': r['distributor__region'] or 'Unknown',
            'revenue': float(r['revenue'] or 0),
            'records': r['records'] or 0,
            'dist_count': r['dist_count'] or 0,
            'color': REGION_COLORS.get(r['distributor__region'], '#6c757d'),
        }
        for r in region_qs
    ]

    # Distributor stats (filtered)
    qs = POSRecord.objects.filter(invoiced_value__isnull=False)
    if selected_region:
        qs = qs.filter(distributor__region=selected_region)
    if selected_sp:
        qs = qs.filter(distributor__salesperson_name=selected_sp)
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)

    dist_data = list(
        _annotate_converted(qs, selected_currency, rates)
        .values('distributor__id', 'distributor__name', 'distributor__region',
                'distributor__salesperson_name', 'distributor__salesperson_code')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'), records=Count('id'))
        .order_by('-revenue')
    )
    total_rev = sum(float(d['revenue'] or 0) for d in dist_data) or 1
    for d in dist_data:
        d['revenue'] = float(d['revenue'] or 0)
        d['units'] = int(d['units'] or 0)
        d['share_pct'] = round(d['revenue'] / total_rev * 100, 1)
        d['color'] = REGION_COLORS.get(d['distributor__region'], '#8205B4')
        # Template-friendly aliases (template uses d.pk, d.name, d.total_revenue, d.total_records)
        d['pk'] = d['distributor__id']
        d['name'] = d['distributor__name']
        d['total_revenue'] = d['revenue']
        d['total_records'] = d['records']
        d['salesperson_name'] = d['distributor__salesperson_name']
        d['salesperson_code'] = d['distributor__salesperson_code']

    top3 = dist_data[:3]

    chart_items = dist_data[:20]
    chart_data = json.dumps({
        'labels': [d['distributor__name'] for d in chart_items],
        'revenue': [d['revenue'] for d in chart_items],
        'colors': [d['color'] for d in chart_items],
        'pks': [d['pk'] for d in chart_items],
    })

    filtered_distributors = None
    if selected_region:
        filtered_distributors = dist_data

    region_info = next((r for r in region_stats if r['name'] == selected_region), None)

    all_salespersons = list(
        Distributor.objects.exclude(salesperson_name='')
        .values_list('salesperson_name', flat=True)
        .distinct()
        .order_by('salesperson_name')
    )

    return render(request, 'reports/distributor_list.html', {
        'all_distributors': all_distributors,
        'region_stats': region_stats,
        'region_info': region_info,
        'selected_region': selected_region,
        'filtered_distributors': filtered_distributors,
        'dist_data': dist_data,
        'top3': top3,
        'chart_data': chart_data,
        'has_filters': bool(date_from or date_to or selected_sp),
        'filters': {'date_from': date_from, 'date_to': date_to, 'region': selected_region, 'salesperson': selected_sp},
        'jump_distributors': jump_distributors,
        'page_title': 'Distributors',
        'selected_currency': selected_currency,
        'currency_symbol': currency_symbol,
        'selected_sp': selected_sp,
        'all_salespersons': all_salespersons,
    })


def product_list(request):
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    search    = request.GET.get('q', '').strip()
    sort_by   = request.GET.get('sort', 'revenue')
    if sort_by not in ('revenue', 'units'):
        sort_by = 'revenue'

    # Region: explicit URL param updates session; navigation falls back to session
    region_param = request.GET.get('region')
    if region_param is not None:
        region = region_param.strip()
        request.session['region'] = region
        request.session.modified = True
    else:
        region = request.session.get('region', '')

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    qs = POSRecord.objects.exclude(manufacturer_part_no='').filter(invoiced_value__isnull=False)
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)
    if search:
        qs = qs.filter(
            Q(manufacturer_part_no__icontains=search) | Q(product_description__icontains=search)
        )
    if region:
        qs = qs.filter(distributor__region=region)

    products = list(
        _annotate_converted(qs, selected_currency, rates)
        .values('manufacturer_part_no')
        .annotate(
            description=Max('product_description'),
            revenue=Sum('converted_value'),
            units=Sum('quantity'),
            dist_count=Count('distributor', distinct=True),
            records=Count('id'),
        )
        .order_by(f'-{sort_by}')
    )
    total_metric = sum(float(p[sort_by] or 0) for p in products) or 1
    for p in products:
        p['revenue'] = float(p['revenue'] or 0)
        p['units'] = int(p['units'] or 0)
        p['share_pct'] = round(float(p[sort_by] or 0) / total_metric * 100, 1)

    all_regions = list(Distributor.objects.filter(region__gt='').values_list('region', flat=True).distinct().order_by('region'))

    return render(request, 'reports/product_list.html', {
        'products': products,
        'filters': {'date_from': date_from, 'date_to': date_to, 'q': search, 'region': region, 'sort': sort_by},
        'has_filters': bool(date_from or date_to or search),
        'sort_by': sort_by,
        'all_regions': all_regions,
        'selected_region': region,
        'page_title': 'Products',
        'selected_currency': selected_currency,
        'currency_symbol': currency_symbol,
    })


def product_detail(request):
    mfr_pn    = request.GET.get('item', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()

    if not mfr_pn:
        return redirect('product_list')

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    qs = POSRecord.objects.filter(manufacturer_part_no=mfr_pn)
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)

    # Prefer Priority canonical English name; fall back to distributor description
    priority_product = PriorityProduct.objects.filter(part_number=mfr_pn).first()
    if priority_product and priority_product.description:
        description = priority_product.description
        priority_family = priority_product.family_description or priority_product.family
        priority_status = priority_product.status
    else:
        first = qs.exclude(product_description='').values('product_description').first()
        description = first['product_description'] if first else ''
        priority_family = ''
        priority_status = ''

    qs_conv = _annotate_converted(qs, selected_currency, rates)
    totals_raw = qs_conv.aggregate(
        revenue=Sum('converted_value'),
        units=Sum('quantity'),
        records=Count('id'),
        customers=Count('customer_name', distinct=True),
    )
    totals = {
        'revenue':   float(totals_raw['revenue'] or 0),
        'units':     int(totals_raw['units'] or 0),
        'records':   int(totals_raw['records'] or 0),
        'customers': int(totals_raw['customers'] or 0),
    }

    by_dist = list(
        qs_conv.values('distributor__id', 'distributor__name', 'distributor__region')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'), records=Count('id'))
        .order_by('-revenue')
    )
    dist_total = sum(float(d['revenue'] or 0) for d in by_dist) or 1
    for d in by_dist:
        d['revenue'] = float(d['revenue'] or 0)
        d['share_pct'] = round(d['revenue'] / dist_total * 100, 1)
        d['color'] = REGION_COLORS.get(d['distributor__region'], '#8205B4')

    top_customers = list(
        qs.exclude(customer_name='')
        .values('customer_name', 'country')
        .annotate(revenue=Sum('invoiced_value'), units=Sum('quantity'), records=Count('id'))
        .order_by('-revenue')[:10]
    )
    for c in top_customers:
        c['revenue'] = float(c['revenue'] or 0)
        c['display_country'] = country_display(c['country'])

    monthly_qs = list(
        _annotate_converted(
            qs.filter(invoice_date__isnull=False, invoiced_value__isnull=False),
            selected_currency, rates
        )
        .annotate(month=TruncMonth('invoice_date'))
        .values('month')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
        .order_by('month')
    )
    chart_data = json.dumps({
        'labels':  [r['month'].strftime('%b %Y') for r in monthly_qs],
        'revenue': [float(r['revenue'] or 0) for r in monthly_qs],
        'units':   [int(r['units'] or 0) for r in monthly_qs],
    })

    return render(request, 'reports/product_detail.html', {
        'mfr_pn':         mfr_pn,
        'description':    description,
        'priority_family': priority_family,
        'priority_status': priority_status,
        'totals':         totals,
        'by_dist':        by_dist,
        'top_customers':  top_customers,
        'chart_data':     chart_data,
        'filters':        {'date_from': date_from, 'date_to': date_to},
        'has_filters':    bool(date_from or date_to),
        'page_title':     description or mfr_pn,
        'selected_currency': selected_currency,
        'currency_symbol':   currency_symbol,
    })


def countries_view(request):
    selected_country = request.GET.get('country', '').strip()
    selected_countries = [selected_country] if selected_country else []
    selected_region = request.GET.get('region', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    qs = POSRecord.objects.exclude(country='').exclude(country__isnull=True)
    if selected_country:
        qs = qs.filter(country=selected_country)
    if selected_region:
        qs = qs.filter(distributor__region=selected_region)
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)

    country_data = list(
        _annotate_converted(qs, selected_currency, rates)
        .values('country')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
        .order_by('-revenue')
    )
    total_rev = sum(float(c['revenue'] or 0) for c in country_data) or 1
    for c in country_data:
        c['revenue'] = float(c['revenue'] or 0)
        c['units'] = int(c['units'] or 0)
        c['share_pct'] = round(c['revenue'] / total_rev * 100, 1)
        c['display_name'] = country_display(c['country'])

    top3 = country_data[:3]
    all_countries = [
        {'code': code, 'name': country_display(code)}
        for code in (
            POSRecord.objects.exclude(country='').exclude(country__isnull=True)
            .values_list('country', flat=True).distinct().order_by('country')
        )
    ]

    chart_items = country_data[:25]
    chart_data = json.dumps({
        'labels': [c['display_name'] for c in chart_items],
        'revenue': [c['revenue'] for c in chart_items],
        'units': [c['units'] for c in chart_items],
    })

    return render(request, 'reports/countries.html', {
        'top3': top3,
        'country_data': country_data,
        'all_countries': all_countries,
        'selected_countries': selected_countries,
        'chart_data': chart_data,
        'has_filters': bool(selected_country or date_from or date_to),
        'selected_region': selected_region,
        'all_regions': Distributor.objects.exclude(region='').values_list('region', flat=True).distinct().order_by('region'),
        'filters': {'date_from': date_from, 'date_to': date_to},
        'page_title': 'Sales by Country',
        'selected_currency': selected_currency,
        'currency_symbol': currency_symbol,
    })


def units_view(request):
    selected_category = request.GET.get('category', '').strip()
    selected_categories = [selected_category] if selected_category else []
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    view_by = request.GET.get('view_by', 'product')

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    qs = POSRecord.objects.filter(quantity__isnull=False)
    if selected_category:
        qs = qs.filter(product_level_1=selected_category)
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)

    if view_by == 'category':
        product_data = list(
            _annotate_converted(
                qs.exclude(product_level_1='').exclude(product_level_1__isnull=True),
                selected_currency, rates
            )
            .values('product_level_1')
            .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
            .order_by('-units')
        )
        for p in product_data:
            p['name'] = p['product_level_1']
    else:
        product_data = list(
            _annotate_converted(
                qs.exclude(product_name='').exclude(product_name__isnull=True),
                selected_currency, rates
            )
            .values('product_name')
            .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
            .order_by('-units')
        )
        for p in product_data:
            p['name'] = p['product_name']

    total_units = sum(int(p['units'] or 0) for p in product_data) or 1
    for p in product_data:
        p['revenue'] = float(p['revenue'] or 0)
        p['units'] = int(p['units'] or 0)
        p['share_pct'] = round(p['units'] / total_units * 100, 1)

    top3 = product_data[:3]
    all_categories = (
        POSRecord.objects.exclude(product_level_1='').exclude(product_level_1__isnull=True)
        .values_list('product_level_1', flat=True).distinct().order_by('product_level_1')
    )

    chart_items = product_data[:20]
    chart_data = json.dumps({
        'labels': [p['name'] for p in chart_items],
        'revenue': [p['revenue'] for p in chart_items],
        'units': [p['units'] for p in chart_items],
    })

    return render(request, 'reports/units.html', {
        'top3': top3,
        'product_data': product_data,
        'all_categories': all_categories,
        'selected_categories': selected_categories,
        'chart_data': chart_data,
        'view_by': view_by,
        'has_filters': bool(selected_category or date_from or date_to),
        'filters': {'date_from': date_from, 'date_to': date_to},
        'page_title': 'Units Sold',
        'selected_currency': selected_currency,
        'currency_symbol': currency_symbol,
    })


def revenue_view(request):
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    qs = POSRecord.objects.filter(invoiced_value__isnull=False)
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)

    qs_conv = _annotate_converted(qs, selected_currency, rates)

    # By region
    region_data = list(
        qs_conv
        .values('distributor__region')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
        .order_by('-revenue')
    )
    for r in region_data:
        r['name'] = r['distributor__region'] or 'Unknown'
        r['revenue'] = float(r['revenue'] or 0)
        r['units'] = int(r['units'] or 0)
        r['color'] = REGION_COLORS.get(r['name'], '#6c757d')
    total_region = sum(r['revenue'] for r in region_data) or 1
    for r in region_data:
        r['share_pct'] = round(r['revenue'] / total_region * 100, 1)

    # By country (top 20)
    country_data = list(
        qs_conv.exclude(country='').values('country')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
        .order_by('-revenue')[:20]
    )
    for c in country_data:
        c['revenue'] = float(c['revenue'] or 0)
        c['units'] = int(c['units'] or 0)
        c['display_name'] = country_display(c['country'])
    total_country = sum(c['revenue'] for c in country_data) or 1
    for c in country_data:
        c['share_pct'] = round(c['revenue'] / total_country * 100, 1)

    # By distributor
    dist_data = list(
        qs_conv
        .values('distributor__id', 'distributor__name', 'distributor__region')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
        .order_by('-revenue')
    )
    for d in dist_data:
        d['revenue'] = float(d['revenue'] or 0)
        d['units'] = int(d['units'] or 0)
        d['color'] = REGION_COLORS.get(d['distributor__region'], '#8205B4')
    total_dist = sum(d['revenue'] for d in dist_data) or 1
    for d in dist_data:
        d['share_pct'] = round(d['revenue'] / total_dist * 100, 1)

    # By product category (top 15)
    product_data = list(
        qs_conv.exclude(product_level_1='').values('product_level_1')
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
        .order_by('-revenue')[:15]
    )
    for p in product_data:
        p['revenue'] = float(p['revenue'] or 0)
        p['units'] = int(p['units'] or 0)
    total_product = sum(p['revenue'] for p in product_data) or 1
    for p in product_data:
        p['share_pct'] = round(p['revenue'] / total_product * 100, 1)

    region_chart = json.dumps({
        'labels': [r['name'] for r in region_data],
        'revenue': [r['revenue'] for r in region_data],
        'colors': [r['color'] for r in region_data],
    })
    country_chart = json.dumps({
        'labels': [c['display_name'] for c in country_data],
        'revenue': [c['revenue'] for c in country_data],
    })
    dist_chart = json.dumps({
        'labels': [d['distributor__name'] for d in dist_data],
        'revenue': [d['revenue'] for d in dist_data],
        'colors': [d['color'] for d in dist_data],
    })
    product_chart = json.dumps({
        'labels': [p['product_level_1'] for p in product_data],
        'revenue': [p['revenue'] for p in product_data],
    })

    total_revenue = sum(r['revenue'] for r in region_data)

    return render(request, 'reports/revenue.html', {
        'region_data': region_data,
        'country_data': country_data,
        'dist_data': dist_data,
        'product_data': product_data,
        'region_chart': region_chart,
        'country_chart': country_chart,
        'dist_chart': dist_chart,
        'product_chart': product_chart,
        'total_revenue': total_revenue,
        'has_filters': bool(date_from or date_to),
        'filters': {'date_from': date_from, 'date_to': date_to},
        'page_title': 'Total Revenue',
        'selected_currency': selected_currency,
        'currency_symbol': currency_symbol,
    })


def salesperson_list(request):
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    selected_sp = request.GET.get('salesperson', '').strip()

    # Region: explicit URL param updates session; navigation falls back to session
    region_param = request.GET.get('region')
    if region_param is not None:
        region = region_param.strip()
        request.session['region'] = region
        request.session.modified = True
    else:
        region = request.session.get('region', '')

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    qs = POSRecord.objects.filter(invoiced_value__isnull=False).exclude(distributor__salesperson_name='')
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)
    if selected_sp:
        qs = qs.filter(distributor__salesperson_name=selected_sp)
    if region:
        qs = qs.filter(distributor__region=region)

    sp_data = list(
        _annotate_converted(qs, selected_currency, rates)
        .values('distributor__salesperson_name', 'distributor__salesperson_code')
        .annotate(
            revenue=Sum('converted_value'),
            units=Sum('quantity'),
            dist_count=Count('distributor', distinct=True),
            records=Count('id'),
        )
        .order_by('-revenue')
    )
    total_rev = sum(float(s['revenue'] or 0) for s in sp_data) or 1
    for s in sp_data:
        s['revenue'] = float(s['revenue'] or 0)
        s['units'] = int(s['units'] or 0)
        s['share_pct'] = round(s['revenue'] / total_rev * 100, 1)
        s['name'] = s['distributor__salesperson_name']
        s['code'] = s['distributor__salesperson_code']

    top3 = sp_data[:3]

    chart_items = sp_data[:10]
    chart_data = json.dumps({
        'labels':  [s['name'] for s in chart_items],
        'revenue': [s['revenue'] for s in chart_items],
    })

    all_salespersons = list(
        Distributor.objects.exclude(salesperson_name='')
        .values_list('salesperson_name', flat=True)
        .distinct()
        .order_by('salesperson_name')
    )
    all_regions = list(Distributor.objects.filter(region__gt='').values_list('region', flat=True).distinct().order_by('region'))

    return render(request, 'reports/salesperson_list.html', {
        'sp_data':         sp_data,
        'top3':            top3,
        'chart_data':      chart_data,
        'all_salespersons': all_salespersons,
        'selected_sp':     selected_sp,
        'all_regions':     all_regions,
        'selected_region': region,
        'has_filters':     bool(date_from or date_to or selected_sp or region),
        'filters':         {'date_from': date_from, 'date_to': date_to, 'region': region},
        'page_title':      'Account Managers',
        'selected_currency': selected_currency,
        'currency_symbol':   currency_symbol,
    })


def weekly_view(request):
    by = request.GET.get('by', 'region')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    selected_currency = request.session.get('currency', 'USD')
    currency_symbol = _currency_symbol(selected_currency)
    rates = _get_rates()

    qs = POSRecord.objects.filter(invoice_date__isnull=False, invoiced_value__isnull=False)
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)

    by_options = {
        'region':      ('distributor__region', 'Region'),
        'distributor': ('distributor__name', 'Distributor'),
        'country':     ('country', 'Country'),
        'product':     ('product_level_1', 'Product Category'),
    }
    group_field, group_label = by_options.get(by, by_options['region'])

    weekly_qs = list(
        _annotate_converted(qs, selected_currency, rates)
        .annotate(week=TruncWeek('invoice_date'))
        .values('week', group_field)
        .annotate(revenue=Sum('converted_value'), units=Sum('quantity'))
        .order_by('week', group_field)
    )

    weeks_sorted = sorted(set(r['week'] for r in weekly_qs))
    week_labels = [w.strftime('W%W — %b %d') for w in weeks_sorted]
    week_keys   = [w.strftime('%Y-%W') for w in weeks_sorted]

    groups = {}
    for r in weekly_qs:
        gval = r[group_field] or 'Unknown'
        if gval not in groups:
            groups[gval] = {}
        groups[gval][r['week'].strftime('%Y-%W')] = float(r['revenue'])

    # Limit to top 10 groups by total revenue
    group_totals = sorted(groups.items(), key=lambda x: -sum(x[1].values()))[:10]

    datasets = []
    for i, (gval, by_week) in enumerate(group_totals):
        if by == 'region':
            color = REGION_COLORS.get(gval, DIST_COLORS[i % len(DIST_COLORS)])
        else:
            color = DIST_COLORS[i % len(DIST_COLORS)]
        datasets.append({
            'label': gval,
            'data': [by_week.get(wk, 0) for wk in week_keys],
            'backgroundColor': color,
            'borderColor': color,
            'borderWidth': 0,
            'fill': False,
            'tension': 0,
            'pointRadius': 0,
            'pointHoverRadius': 0,
            'borderRadius': 4,
            '_color': color,
        })

    chart_data = json.dumps({'labels': week_labels, 'datasets': datasets})

    return render(request, 'reports/weekly.html', {
        'chart_data': chart_data,
        'by': by,
        'group_label': group_label,
        'has_filters': bool(date_from or date_to),
        'filters': {'date_from': date_from, 'date_to': date_to},
        'page_title': 'Sales by Week',
        'selected_currency': selected_currency,
        'currency_symbol': currency_symbol,
    })


# ── AI Assistant ──────────────────────────────────────────────────────────────

def _build_ai_context():
    """Minimal metadata for the AI system prompt (date range + distributor list). Cached 5 min."""
    cached = cache.get('kpos_ai_context')
    if cached:
        return cached

    dates = POSRecord.objects.aggregate(min_date=Min('invoice_date'), max_date=Max('invoice_date'))
    distributors = list(Distributor.objects.values('name', 'region', 'code').order_by('region', 'name'))
    lines = [
        f"Database date range: {dates['min_date']} to {dates['max_date']}",
        f"\nDistributors ({len(distributors)} total — use code for filtering):",
    ]
    for d in distributors:
        lines.append(f"  {d['name']} | region: {d['region']} | code: {d['code']}")
    context = '\n'.join(lines)
    cache.set('kpos_ai_context', context, 300)
    return context


# ── Tool definitions ───────────────────────────────────────────────────────────

KPOS_TOOLS = [
    {
        "name": "get_top_products",
        "description": "Get top-selling products ranked by revenue or units. Filter by any combination of time period, distributor, or region.",
        "input_schema": {
            "type": "object",
            "properties": {
                "year":             {"type": "integer", "description": "Filter by year (e.g. 2026)"},
                "month":            {"type": "integer", "description": "Filter by month number 1–12"},
                "quarter":          {"type": "integer", "description": "Filter by quarter 1–4"},
                "distributor_code": {"type": "string",  "description": "Distributor code (e.g. 'cdev')"},
                "region":           {"type": "string",  "description": "Region name (e.g. 'EMEA')"},
                "sort_by":          {"type": "string",  "enum": ["revenue", "units"]},
                "limit":            {"type": "integer", "description": "Max results, default 10"},
            },
        },
    },
    {
        "name": "get_top_distributors",
        "description": "Get top distributors ranked by revenue or units. Filter by time period or region.",
        "input_schema": {
            "type": "object",
            "properties": {
                "year":    {"type": "integer"},
                "month":   {"type": "integer"},
                "quarter": {"type": "integer"},
                "region":  {"type": "string"},
                "sort_by": {"type": "string", "enum": ["revenue", "units"]},
                "limit":   {"type": "integer"},
            },
        },
    },
    {
        "name": "get_top_customers",
        "description": "Get top customers ranked by revenue or units. Filter by time period, distributor, region, or country.",
        "input_schema": {
            "type": "object",
            "properties": {
                "year":             {"type": "integer"},
                "month":            {"type": "integer"},
                "quarter":          {"type": "integer"},
                "distributor_code": {"type": "string"},
                "region":           {"type": "string"},
                "country":          {"type": "string", "description": "Country name (partial match)"},
                "sort_by":          {"type": "string", "enum": ["revenue", "units"]},
                "limit":            {"type": "integer"},
            },
        },
    },
    {
        "name": "get_top_sales_reps",
        "description": "Get sales representative performance ranked by revenue. Filter by time period, distributor, or region.",
        "input_schema": {
            "type": "object",
            "properties": {
                "year":             {"type": "integer"},
                "month":            {"type": "integer"},
                "quarter":          {"type": "integer"},
                "distributor_code": {"type": "string"},
                "region":           {"type": "string"},
                "limit":            {"type": "integer"},
            },
        },
    },
    {
        "name": "get_revenue_trend",
        "description": "Get month-by-month revenue and units trend. Filter by year, distributor, region, or product.",
        "input_schema": {
            "type": "object",
            "properties": {
                "year":             {"type": "integer"},
                "distributor_code": {"type": "string"},
                "region":           {"type": "string"},
                "product_name":     {"type": "string", "description": "Product name partial match"},
            },
        },
    },
    {
        "name": "get_summary",
        "description": "Get KPI summary: total revenue, units, customer count, product count. Filter by any time period, distributor, or region.",
        "input_schema": {
            "type": "object",
            "properties": {
                "year":             {"type": "integer"},
                "month":            {"type": "integer"},
                "quarter":          {"type": "integer"},
                "distributor_code": {"type": "string"},
                "region":           {"type": "string"},
            },
        },
    },
]


def _apply_filters(qs, year=None, month=None, quarter=None,
                   distributor_code=None, region=None, country=None, product_name=None):
    if year:
        qs = qs.filter(invoice_date__year=int(year))
    if month:
        qs = qs.filter(invoice_date__month=int(month))
    if quarter:
        q_months = {1: [1, 2, 3], 2: [4, 5, 6], 3: [7, 8, 9], 4: [10, 11, 12]}
        qs = qs.filter(invoice_date__month__in=q_months.get(int(quarter), []))
    if distributor_code:
        qs = qs.filter(distributor__code__iexact=str(distributor_code))
    if region:
        qs = qs.filter(distributor__region__iexact=str(region))
    if country:
        qs = qs.filter(country__icontains=str(country))
    if product_name:
        qs = qs.filter(
            Q(manufacturer_part_no__icontains=str(product_name)) |
            Q(product_description__icontains=str(product_name))
        )
    return qs


def _tool_get_top_products(year=None, month=None, quarter=None,
                           distributor_code=None, region=None, sort_by='revenue', limit=10):
    limit = min(int(limit or 10), 30)
    qs = _apply_filters(_annotate_converted(POSRecord.objects.all(), 'USD'),
                        year=year, month=month, quarter=quarter,
                        distributor_code=distributor_code, region=region)
    rows = (qs.exclude(manufacturer_part_no='')
            .values('manufacturer_part_no', 'product_description')
            .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'))
            .order_by('-total_usd' if sort_by != 'units' else '-total_qty')[:limit])
    if not rows:
        return "No product data found for those filters."
    return '\n'.join(
        f"{i}. {r['manufacturer_part_no']} ({r['product_description'] or 'no description'}): "
        f"${r['total_usd'] or 0:,.0f} | {r['total_qty'] or 0:,} units"
        for i, r in enumerate(rows, 1)
    )


def _tool_get_top_distributors(year=None, month=None, quarter=None,
                               region=None, sort_by='revenue', limit=10):
    limit = min(int(limit or 10), 30)
    qs = _apply_filters(_annotate_converted(POSRecord.objects.all(), 'USD'),
                        year=year, month=month, quarter=quarter, region=region)
    rows = (qs.values('distributor__name', 'distributor__region')
            .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'))
            .order_by('-total_usd' if sort_by != 'units' else '-total_qty')[:limit])
    if not rows:
        return "No distributor data found for those filters."
    return '\n'.join(
        f"{i}. {r['distributor__name']} ({r['distributor__region']}): "
        f"${r['total_usd'] or 0:,.0f} | {r['total_qty'] or 0:,} units"
        for i, r in enumerate(rows, 1)
    )


def _tool_get_top_customers(year=None, month=None, quarter=None,
                            distributor_code=None, region=None, country=None, sort_by='revenue', limit=10):
    limit = min(int(limit or 10), 30)
    qs = _apply_filters(_annotate_converted(POSRecord.objects.all(), 'USD'),
                        year=year, month=month, quarter=quarter,
                        distributor_code=distributor_code, region=region, country=country)
    rows = (qs.exclude(customer_name='')
            .values('customer_name', 'country', 'distributor__region')
            .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'))
            .order_by('-total_usd' if sort_by != 'units' else '-total_qty')[:limit])
    if not rows:
        return "No customer data found for those filters."
    return '\n'.join(
        f"{i}. {r['customer_name']} ({r['country'] or '—'}, {r['distributor__region']}): "
        f"${r['total_usd'] or 0:,.0f} | {r['total_qty'] or 0:,} units"
        for i, r in enumerate(rows, 1)
    )


def _tool_get_top_sales_reps(year=None, month=None, quarter=None,
                             distributor_code=None, region=None, limit=10):
    limit = min(int(limit or 10), 30)
    qs = _apply_filters(_annotate_converted(POSRecord.objects.all(), 'USD'),
                        year=year, month=month, quarter=quarter,
                        distributor_code=distributor_code, region=region)
    rows = (qs.filter(distributor__salesperson_name__gt='')
            .values('distributor__salesperson_name', 'distributor__region')
            .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'))
            .order_by('-total_usd')[:limit])
    if not rows:
        return "No sales rep data found for those filters."
    return '\n'.join(
        f"{i}. {r['distributor__salesperson_name']} ({r['distributor__region']}): "
        f"${r['total_usd'] or 0:,.0f} | {r['total_qty'] or 0:,} units"
        for i, r in enumerate(rows, 1)
    )


def _tool_get_revenue_trend(year=None, distributor_code=None, region=None, product_name=None):
    qs = _apply_filters(_annotate_converted(POSRecord.objects.all(), 'USD'),
                        year=year, distributor_code=distributor_code,
                        region=region, product_name=product_name)
    rows = (qs.annotate(month=TruncMonth('invoice_date'))
            .values('month')
            .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'))
            .order_by('month'))
    if not rows:
        return "No trend data found for those filters."
    lines = ["Month | Revenue (USD) | Units"]
    for r in rows:
        if r['month']:
            lines.append(f"{r['month'].strftime('%Y-%m')}: ${r['total_usd'] or 0:,.0f} | {r['total_qty'] or 0:,} units")
    return '\n'.join(lines)


def _tool_get_summary(year=None, month=None, quarter=None, distributor_code=None, region=None):
    qs = _apply_filters(_annotate_converted(POSRecord.objects.all(), 'USD'),
                        year=year, month=month, quarter=quarter,
                        distributor_code=distributor_code, region=region)
    agg = qs.aggregate(
        total_usd=Sum('converted_value'), total_qty=Sum('quantity'),
        record_count=Count('id'),
        customer_count=Count('customer_name', distinct=True),
        distributor_count=Count('distributor', distinct=True),
        product_count=Count('manufacturer_part_no', distinct=True),
    )
    dates = qs.aggregate(min_date=Min('invoice_date'), max_date=Max('invoice_date'))
    return '\n'.join([
        f"Total revenue (USD): ${agg['total_usd'] or 0:,.0f}",
        f"Total units sold:    {agg['total_qty'] or 0:,}",
        f"Invoice records:     {agg['record_count']:,}",
        f"Unique customers:    {agg['customer_count']:,}",
        f"Distributors:        {agg['distributor_count']:,}",
        f"Unique products:     {agg['product_count']:,}",
        f"Date range:          {dates['min_date']} to {dates['max_date']}",
    ])


def _execute_tool(name, inputs):
    dispatch = {
        'get_top_products':     _tool_get_top_products,
        'get_top_distributors': _tool_get_top_distributors,
        'get_top_customers':    _tool_get_top_customers,
        'get_top_sales_reps':   _tool_get_top_sales_reps,
        'get_revenue_trend':    _tool_get_revenue_trend,
        'get_summary':          _tool_get_summary,
    }
    try:
        fn = dispatch.get(name)
        if not fn:
            return f"Unknown tool: {name}"
        return fn(**{k: v for k, v in inputs.items() if v is not None})
    except Exception as exc:
        return f"Tool error: {exc}"


def ai_chat(request):
    """AI assistant — agentic tool-use loop against live DB."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        body = json.loads(request.body)
    except (ValueError, KeyError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_message = (body.get('message') or '').strip()
    if not user_message:
        return JsonResponse({'error': 'Empty message'}, status=400)

    api_key = getattr(django_settings, 'CLAUDE_API_KEY', '')
    if not api_key:
        return JsonResponse({'reply': 'AI assistant is not configured yet.', 'wants_export': False})

    data_context = _build_ai_context()
    today_str = date_cls.today().strftime('%B %d, %Y')

    system_prompt = f"""You are KPOS Assistant, an AI analyst built into Kramer Electronics' KPOS Point-of-Sale analytics platform.

Use the provided tools to query live data. Always call a tool rather than guessing or relying on memory.

{data_context}

Today: {today_str}. All revenue figures are normalized to USD.

Guidelines:
- Keep replies short and direct — 1–3 sentences unless detail is genuinely needed. No preamble or filler.
- Use conversation context for follow-ups: if the user asks "which are the top 3?" after asking about February, query for February.
- Format numbers with commas and $ prefix (e.g., $1,234,567).
- If asked to export to Excel or spreadsheet, end your reply with exactly: [EXPORT_EXCEL]
"""

    history = request.session.get('ai_chat_history', [])
    history.append({'role': 'user', 'content': user_message})

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        reply = ''

        for _ in range(6):  # max tool-use rounds
            response = client.messages.create(
                model='claude-opus-4-8',
                max_tokens=1024,
                system=system_prompt,
                messages=history,
                tools=KPOS_TOOLS,
            )

            if response.stop_reason == 'tool_use':
                # Serialize assistant content (text + tool_use blocks) for session storage
                asst_content = []
                for blk in response.content:
                    if blk.type == 'text':
                        asst_content.append({'type': 'text', 'text': blk.text})
                    elif blk.type == 'tool_use':
                        asst_content.append({
                            'type': 'tool_use',
                            'id': blk.id,
                            'name': blk.name,
                            'input': blk.input,
                        })
                history.append({'role': 'assistant', 'content': asst_content})

                # Execute each tool and collect results
                tool_results = []
                for blk in response.content:
                    if blk.type == 'tool_use':
                        result = _execute_tool(blk.name, blk.input)
                        tool_results.append({
                            'type': 'tool_result',
                            'tool_use_id': blk.id,
                            'content': result,
                        })
                history.append({'role': 'user', 'content': tool_results})

            else:  # end_turn
                reply = ''.join(blk.text for blk in response.content if blk.type == 'text')
                break

        if not reply:
            reply = "I wasn't able to complete that request. Please try again."

    except Exception as exc:
        return JsonResponse({'error': f'AI error: {exc}'}, status=500)

    wants_export = '[EXPORT_EXCEL]' in reply
    clean_reply = reply.replace('[EXPORT_EXCEL]', '').strip()

    history.append({'role': 'assistant', 'content': clean_reply})
    if len(history) > 40:
        history = history[-40:]

    request.session['ai_chat_history'] = history
    request.session.modified = True

    return JsonResponse({'reply': clean_reply, 'wants_export': wants_export})


def ai_export(request):
    """Generate a comprehensive Excel export for the AI assistant download button."""
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    hdr_fill = PatternFill(start_color='8200B4', end_color='8200B4', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    hdr_align = Alignment(horizontal='center')

    def make_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align

    def autofit(ws):
        for col in ws.columns:
            width = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 50)

    qs = _annotate_converted(POSRecord.objects.all(), 'USD')

    ws1 = wb.active
    ws1.title = 'Distributor Summary'
    make_header(ws1, ['Distributor', 'Region', 'Revenue (USD)', 'Units', 'Records'])
    for d in (
        qs.values('distributor__name', 'distributor__region')
        .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'), record_count=Count('id'))
        .order_by('-total_usd')
    ):
        ws1.append([
            d['distributor__name'], d['distributor__region'],
            float(d['total_usd'] or 0), int(d['total_qty'] or 0), d['record_count'],
        ])
    autofit(ws1)

    ws2 = wb.create_sheet('Product Summary')
    make_header(ws2, ['Part No', 'Description', 'Revenue (USD)', 'Units', 'Distributors'])
    for p in (
        qs.exclude(manufacturer_part_no='')
        .values('manufacturer_part_no', 'product_description')
        .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'), dist_count=Count('distributor', distinct=True))
        .order_by('-total_usd')[:100]
    ):
        ws2.append([
            p['manufacturer_part_no'], p['product_description'] or '',
            float(p['total_usd'] or 0), int(p['total_qty'] or 0), p['dist_count'],
        ])
    autofit(ws2)

    ws3 = wb.create_sheet('Monthly Trend')
    make_header(ws3, ['Month', 'Revenue (USD)', 'Units'])
    for m in (
        qs.annotate(month=TruncMonth('invoice_date'))
        .values('month')
        .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'))
        .order_by('month')
    ):
        if m['month']:
            ws3.append([m['month'].strftime('%Y-%m'), float(m['total_usd'] or 0), int(m['total_qty'] or 0)])
    autofit(ws3)

    ws4 = wb.create_sheet('Customer Summary')
    make_header(ws4, ['Customer', 'Country', 'Region', 'Revenue (USD)', 'Units'])
    for c in (
        qs.exclude(customer_name='')
        .values('customer_name', 'country', 'distributor__region')
        .annotate(total_usd=Sum('converted_value'), total_qty=Sum('quantity'))
        .order_by('-total_usd')[:200]
    ):
        ws4.append([
            c['customer_name'], c['country'] or '', c['distributor__region'] or '',
            float(c['total_usd'] or 0), int(c['total_qty'] or 0),
        ])
    autofit(ws4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    today_str = date_cls.today().strftime('%Y%m%d')
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="KPOS_Export_{today_str}.xlsx"'
    return response
